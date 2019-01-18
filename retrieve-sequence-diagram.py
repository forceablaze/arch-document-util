#!/usr/bin/env python
#-*- coding: utf-8 -*-

from __future__ import print_function


from optparse import OptionParser

import io, sys, os
import csv
import re
import openpyxl as pyxl

"""
from cvlog.symbol import SymbolTable
from cvlog.symbol import Symbol
from cvlog.exceptions import NoSymbolFoundException
"""

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.worksheet.write_only import WriteOnlyCell

from pathlib import Path
from utils import isSequenceMatchPattern

# report title color
REPORT_TITLE_COLOR = 'B7DEE8'

# blue
TABLE_TITLE_COLOR = 'FF00B0F0'

# yellow
ITEM_HEADER_COLOR = 'FFFFFF00'

# ID color theme value
ID_COLOR_VALUE = 0.3999755851924192

# except value theme value
EXCEPT_COLOR_VALUE = 0.7999816888943144


ID_DELETED_COLOR = -0.249977111117893


titleStyle = NamedStyle(name="titleStyle")
titleStyle.font = Font(bold=True)
titleStyle.fill = PatternFill("solid", fgColor=REPORT_TITLE_COLOR)
titleBorder = Side("thin", color='000000')
titleStyle.border = Border(left=titleBorder, top=titleBorder, right=titleBorder, bottom=titleBorder)

def printHelpMessageAndExit(option, parser):
    print(option, 'option is required.')
    parser.print_help()
    sys.exit(1)

def style_range(ws, cell_range, style):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    """
    rows = ws[cell_range]

    for row in rows:
        l = row[0]
        r = row[-1]
        if style:
            for c in row:
                c.style = style


def generateReport(resultDict, reportPath):
    report = pyxl.Workbook()

    ws = report.create_sheet(title = u'報告情報')
    ws.append([u'機能名', u'ID総数', u'TBD総数', u'TBD ID詳細', u'NONE数'])

    for fileName in sorted(resultDict):
        data = []
        data.append('{}'.format(Path(fileName).stem))

        idCount = 0
        tbdCount = 0
        tbdIdStr = None
        noneCount = 0

        # no sequence diagram
        if resultDict[fileName] is None:
            continue

        for tableName, table in resultDict[fileName].items():

            for key, val in table.items():
                # ID or ID+NONE, ...
                idCount += val['count']

                # TODO find symbol
                """
                for i in range(0, len(val['items'])):
                    if val['items'][i][5] is not None:
                        value = int(val['items'][i][5])
                        print('ID', value)
                        try:
                            symbol = symbolTable.getSymbol('apitype', value)
                            print(symbol.getSymbolName())
                        except NoSymbolFoundException as e:
                            print(e.message)
                            pass

                    else:
                        pass
                """

                if key == 'ID+TBD':
                    tbdCount += val['count']
                    idListStr = io.StringIO()

                    # item row data
                    for i in range(0, len(val['items'])):
                        idListStr.write(u'{}'.format(val['items'][i][5]))

                        if i < len(val['items']) - 1:
                            idListStr.write(u',')

                    tbdIdStr = idListStr.getvalue()
                    idListStr.close()

                elif key == 'ID+NONE':
                    noneCount += val['count']

        data.append(idCount)
        data.append(tbdCount)
        data.append(tbdIdStr)
        data.append(noneCount)

        ws.append(data)

    style_range(ws, 'A1:E1', style = titleStyle)

    report.save(reportPath)


def showRowItem(titleRow, dataRow):

    for item in titleRow:
        print(item.value, ' ', end = '')
    print('')

    for item in dataRow:
        print(item.value, ' ', end = '')
    print('')

def checkAndCreateDict(header, item, tableDict):
    if header[5].value not in tableDict:
        tableDict[header[5].value] = {}
        tableDict[header[5].value]['count'] = 0
        tableDict[header[5].value]['none_count'] = 0
        tableDict[header[5].value]['items'] = []


    # this id is not enable
    #if item[1].fill.start_color.tint == ID_DELETED_COLOR:
    #    return

    if item[5].value is not None:
        itemArray = []
        for cell in item:
            itemArray.append(u'{}'.format(cell.value))

        tableDict[header[5].value]['count'] += 1
        tableDict[header[5].value]['items'].append(itemArray)
    else:
        tableDict[header[5].value]['none_count'] += 1

def readFeatureTable(rowIter, tableDict, titleRow = None):

    if isSequenceMatchPattern(u'期待値', u'{}'.format(titleRow[5].value)):
        print('retrieve next row')
        header = rowIter.next()
        item = rowIter.next()
        checkAndCreateDict(header, item, tableDict)


    for row in rowIter:
        if row[1] is not None and row[1].fill is not None:
            if row[1].fill.start_color.rgb == ITEM_HEADER_COLOR:
                print('header')

            # meet next table
            elif row[1].fill.start_color.rgb == TABLE_TITLE_COLOR:
                print(u'Found next table {}'.format(row[1].value))
                return row

        if row[5] is not None:
            #if row[5].fill.start_color.type == 'theme':
                #if row[5].fill.start_color.tint == ID_COLOR_VALUE:
            if isSequenceMatchPattern(u'期待値', u'{}'.format(row[5].value)):
                print('retrieve next row')
                header = rowIter.next()
                item = rowIter.next()
                    #showRowItem(row, item)
                checkAndCreateDict(header, item, tableDict)

def checkTableType(row, rowIter, resultDict):

    if row is None:
        """
        for key, val in resultDict.items():
            # ID or ID+NONE, ...
            print(u'{}'.format(key))

            # item row data
            for item in val['items']:
                print(u'{}, '.format(item[5]), end = '')
            print('')
        """
        return resultDict

    #if(isSequenceMatchPattern(u'機能', u'{}'.format(row[1].value))):
    if row[1] is not None and row[1].fill is not None:
        if row[1].fill.start_color.rgb == TABLE_TITLE_COLOR:
            print(u'Found {} table {}'.format(row[1].value, row[2].value))

            if row[1].value not in resultDict:
                print('insert new table')
                resultDict[row[1].value] = {}

            try:
                currentRow = readFeatureTable(rowIter, resultDict[row[1].value], titleRow = row)
            except StopIteration:
                return resultDict

            checkTableType(currentRow, rowIter, resultDict)

def parseSequenceSheet(sheet, resultDict):
    rowIter = sheet.iter_rows()
    for row in rowIter:
        checkTableType(row, rowIter, resultDict)

def handleXLSMDocument(docPath):
    try:
        wb = pyxl.load_workbook(str(docPath), read_only=True);
    except IOError as e:
        print(e)
        return;

    for name in wb.sheetnames:
        if(isSequenceMatchPattern(u'ドメイン間シーケンス図', name)):
            print(u'Sheet match the pattern, parse it {}'.format(name))
            resultDict = {}
            parseSequenceSheet(wb[name], resultDict)
            return resultDict

    return None

def searchDocument(folder, suffix):
    for root, dirs, files in os.walk(folder):
        for _file in files:
            fPath = Path(root, _file)
            if fPath.suffix != suffix:
                continue
            print('handle document {}'.format(fPath))

            result = handleXLSMDocument(fPath)
            retrievedResult[fPath] = result

if __name__ == '__main__':
    parser = OptionParser()

    # options settings
    parser.add_option("-F", "--folder", default=None,
        action="store", dest = "folder",
        help = u"設計書 folder")

    (options, args) = parser.parse_args()


    for option, value in vars(options).items():
        if value is None:
            printHelpMessageAndExit(option, parser)

    """
    symbolTable = SymbolTable.loadSymbolTable(options.symbol)
    apitypeSymbols = symbolTable.getSymbolsByTag('apitype')
    """

    retrievedResult = {}
    documentList = searchDocument(options.folder, ".xlsm")

    for fileName, result in retrievedResult.items():
        print('{}'.format(fileName))

        idCount = 0

        # no sequence diagram
        if result is None:
            continue

        for tableName, table in result.items():
            print(u'\t{}'.format(tableName))


            for key, val in table.items():
                # ID or ID+NONE, ...
                print(u'\t\t{} {} {}'.format(key, val['count'], val['none_count']))
                idCount += val['count']

                # item row data
                for item in val['items']:
                    print(u'\t\t{}, '.format(item[5]), end = '')
                print('')
    generateReport(retrievedResult, 'result.xlsx')
